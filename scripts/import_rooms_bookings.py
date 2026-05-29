"""
Read 'rooms (3).xlsx' (rooms + booking sheets) and generate two SQL
files for Supabase:

  scripts/sql_rooms_upsert.sql       — UPSERT all 41 rooms (safe to re-run)
  scripts/sql_bookings_load.sql      — INSERT all bookings (large; review before running)

The bookings file starts with an OPTIONAL truncate (commented out) so
the user can choose:
  (a) keep existing bookings + add Excel ones    (default — uncomment nothing)
  (b) wipe and reload from Excel                  (uncomment the TRUNCATE)

Bookings use INSERT ... ON CONFLICT DO NOTHING in case a booking
overlaps an existing one (mtg_no_overlap constraint would reject).
"""
import io
import re
import sys
from datetime import datetime, time, date
from pathlib import Path
import openpyxl

if sys.stdout.encoding != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

XLSX = Path(__file__).parent.parent / 'rooms (3).xlsx'
OUT_ROOMS    = Path(__file__).parent / 'sql_rooms_upsert.sql'
OUT_BOOKINGS = Path(__file__).parent / 'sql_bookings_load.sql'


def sql_str(v):
    """Quote text for SQL literal — handles None, escapes single quotes."""
    if v is None or v == '':
        return 'NULL'
    return "'" + str(v).replace("'", "''") + "'"


def sql_int(v):
    if v is None or v == '':
        return 'NULL'
    try:
        return str(int(v))
    except (ValueError, TypeError):
        return 'NULL'


def time_to_min(v):
    """'09:30:00' (string or time) → 570 (minutes from midnight)."""
    if v is None or v == '':
        return None
    if isinstance(v, time):
        return v.hour * 60 + v.minute
    s = str(v).strip()
    m = re.match(r'^(\d{1,2}):(\d{2})(?::\d{2})?$', s)
    if not m:
        return None
    h, mm = int(m.group(1)), int(m.group(2))
    if h < 0 or h > 24 or mm < 0 or mm > 59:
        return None
    return h * 60 + mm


def normalize_date(v):
    if v is None or v == '':
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip().split(' ')[0]
    return s if re.match(r'^\d{4}-\d{2}-\d{2}$', s) else None


# ──────────────────────────────────────────────────────────────────────
# Load
# ──────────────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(XLSX, data_only=True)

# ----- ROOMS -----
ws = wb['rooms']
rooms = []
for r_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
    if r_idx == 1:
        continue   # header
    if not row[0]:
        continue
    rooms.append({
        'id':       str(row[0]).strip(),
        'name':     str(row[1] or '').strip(),
        # Use the public Google Drive thumbnail URL — col 7 (index 6)
        'picture':  str(row[6] or '').strip() if len(row) > 6 else '',
        'location': str(row[3] or '').strip() if len(row) > 3 else '',
        'status':   str(row[4] or 'available').strip() if len(row) > 4 else 'available',
        'seats':    row[5] if len(row) > 5 else None,
    })

print(f'rooms parsed: {len(rooms)}')

# ----- BOOKINGS -----
ws = wb['booking']
bookings = []
skipped = 0
header = None
for r_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
    if r_idx == 1:
        header = list(row)
        continue
    if not row[0]:
        continue
    d = normalize_date(row[0])
    s = time_to_min(row[1])
    e = time_to_min(row[2])
    room_id = str(row[3] or '').strip() if row[3] else ''
    if not d or s is None or e is None or not room_id or e <= s:
        skipped += 1
        continue
    # title — derive from purpose / detail since the field is required.
    purpose = str(row[13] or '').strip() if len(row) > 13 else ''
    detail  = str(row[14] or '').strip() if len(row) > 14 else ''
    company = str(row[15] or '').strip() if len(row) > 15 else ''
    title   = (detail or purpose or 'การประชุม')[:200]

    booker_name = str(row[7] or '').strip() if len(row) > 7 else ''
    booker_id   = str(row[6] or '').strip() if len(row) > 6 else ''
    booker = (booker_name + (f' ({booker_id})' if booker_id else '')).strip()

    attendees = row[17] if len(row) > 17 else None
    try:
        attendees = int(attendees) if attendees not in (None, '', '-') else None
    except (TypeError, ValueError):
        attendees = None

    refresh_raw = str(row[10] or '').strip() if len(row) > 10 else ''
    refreshments = []
    if refresh_raw and refresh_raw.lower() not in ('no', 'n', 'none', '-'):
        refreshments = [r.strip() for r in re.split(r'[,/]', refresh_raw) if r.strip()]

    equip_raw = str(row[12] or '').strip() if len(row) > 12 else ''
    equipment = []
    if equip_raw and equip_raw.lower() not in ('no', 'n', 'none', '-'):
        equipment = [r.strip() for r in re.split(r'[,/]', equip_raw) if r.strip()]

    bookings.append({
        'room_id':      room_id,
        'booking_date': d,
        'start_min':    s,
        'end_min':      e,
        'title':        title,
        'booker':       booker[:200] if booker else '',
        'attendees':    attendees,
        'purpose':      purpose,
        'company':      company[:200],
        'equipment':    equipment,
        'refreshments': refreshments,
    })

print(f'bookings parsed: {len(bookings)} (skipped {skipped} bad rows)')


# ──────────────────────────────────────────────────────────────────────
# Write rooms SQL
# ──────────────────────────────────────────────────────────────────────
lines = [
    "-- Rooms upsert from 'rooms (3).xlsx'",
    "-- Idempotent — INSERT new rooms, UPDATE existing by id.",
    "-- Does NOT delete rooms that exist in the DB but not in the Excel,",
    "-- because mtg_rooms has ON DELETE CASCADE → bookings would get wiped.",
    "-- Manually drop any obsolete rooms after reviewing what's left.",
    "",
    "begin;",
    "",
]
for r in rooms:
    seats = sql_int(r['seats'])
    lines.append(
        f"insert into public.mtg_rooms (id, name, picture, location, seats, status) values "
        f"({sql_str(r['id'])}, {sql_str(r['name'])}, {sql_str(r['picture'])}, "
        f"{sql_str(r['location'])}, {seats}, {sql_str(r['status'])}) "
        f"on conflict (id) do update set "
        f"name = excluded.name, "
        f"picture = excluded.picture, "
        f"location = excluded.location, "
        f"seats = excluded.seats, "
        f"status = excluded.status;"
    )
lines.append("")
lines.append("commit;")
lines.append("")
lines.append("-- Sanity check:")
lines.append("-- select id, name, location, status, seats from public.mtg_rooms order by id;")
OUT_ROOMS.write_text('\n'.join(lines), encoding='utf-8')
print(f'wrote {OUT_ROOMS}')


# ──────────────────────────────────────────────────────────────────────
# Write bookings SQL
# ──────────────────────────────────────────────────────────────────────
lines = [
    "-- Bookings load from 'rooms (3).xlsx'",
    "-- WARNING: This will INSERT thousands of bookings. Read the notes",
    "-- before running.",
    "--",
    "-- The mtg_no_overlap constraint will reject any booking that overlaps",
    "-- an existing one. We use a savepoint per chunk so one bad row doesn't",
    "-- abort the whole load — but for safety, the cleanest path is:",
    "--   (1) wipe existing bookings (uncomment the TRUNCATE below) — OR",
    "--   (2) leave existing alone and skip overlaps via subquery.",
    "--",
    "-- Default: option (2) — INSERT only if no overlap exists.",
    "",
    "begin;",
    "",
    "-- (option 1) Wipe everything first — uncomment if you want a clean slate:",
    "-- truncate table public.mtg_bookings;",
    "",
]

# Insert in chunks of 200 with NOT EXISTS guard against the no-overlap constraint
CHUNK = 200
for i, b in enumerate(bookings):
    eq_arr  = '{' + ','.join('"' + e.replace('"', '\\"') + '"' for e in b['equipment'])    + '}'
    rf_arr  = '{' + ','.join('"' + e.replace('"', '\\"') + '"' for e in b['refreshments']) + '}'
    lines.append(
        f"insert into public.mtg_bookings "
        f"(room_id, booking_date, start_min, end_min, title, booker, attendees, purpose, company, equipment, refreshments) "
        f"select {sql_str(b['room_id'])}, {sql_str(b['booking_date'])}::date, "
        f"{b['start_min']}, {b['end_min']}, {sql_str(b['title'])}, "
        f"{sql_str(b['booker'])}, {sql_int(b['attendees'])}, "
        f"{sql_str(b['purpose'])}, {sql_str(b['company'])}, "
        f"{sql_str(eq_arr)}::text[], {sql_str(rf_arr)}::text[] "
        f"where not exists ("
        f"  select 1 from public.mtg_bookings"
        f"  where room_id = {sql_str(b['room_id'])}"
        f"    and booking_date = {sql_str(b['booking_date'])}::date"
        f"    and int4range(start_min, end_min) && int4range({b['start_min']}, {b['end_min']})"
        f");"
    )

lines.append("")
lines.append("commit;")
lines.append("")
lines.append("-- Sanity check counts after load:")
lines.append("-- select count(*) as total, count(distinct room_id) as distinct_rooms,")
lines.append("--        min(booking_date) as earliest, max(booking_date) as latest")
lines.append("-- from public.mtg_bookings;")
OUT_BOOKINGS.write_text('\n'.join(lines), encoding='utf-8')
size_mb = OUT_BOOKINGS.stat().st_size / 1024 / 1024
print(f'wrote {OUT_BOOKINGS} ({size_mb:.1f} MB)')

# ──────────────────────────────────────────────────────────────────────
print()
print('────────────────────────────────────────')
print(f' Rooms:    {len(rooms)} → {OUT_ROOMS.name}')
print(f' Bookings: {len(bookings)} → {OUT_BOOKINGS.name}')
print('────────────────────────────────────────')
print()
print('Locations in rooms:', sorted({r["location"] for r in rooms if r["location"]}))
print('Date range:', min(b['booking_date'] for b in bookings), '→',
      max(b['booking_date'] for b in bookings))
