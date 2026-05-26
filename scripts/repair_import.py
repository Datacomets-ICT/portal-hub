"""
Import Repair (ช่าง) historical data from
'Repair/ใบแจ้งซ่อม.xlsx' → SQL INSERTs for rpr_jobs + rpr_inspections.

The xlsx is exported from the legacy PowerApps system. Two sheets matter
for Phase 1:
  - 'ใบแจ้งซ่อม' (922 rows, 47 cols) → rpr_jobs
  - 'ใบตรวจประจำเดือน' (1000 rows, 33 cols) → rpr_inspections

Image paths in the xlsx are PowerApps internal references like
"ใบแจ้งซ่อม_Images/cb3d3cf...". They don't exist in our Supabase
storage, so photo_urls are left empty here — backfilled later if/when
the original image files are recovered.

Output: scripts/repair_import.sql (idempotent; uses ON CONFLICT DO NOTHING
so re-running won't duplicate). Run in Supabase SQL Editor after
schema_rpr_v1.sql is applied.

Run: python scripts/repair_import.py
"""
import io
import sys
from pathlib import Path
import openpyxl

if sys.stdout.encoding != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

XLSX = Path(__file__).parent.parent / 'Repair' / 'ใบแจ้งซ่อม.xlsx'
OUT  = Path(__file__).parent / 'repair_import.sql'

# ────────────────────────────────────────────────────────────────────
# Helpers
# ────────────────────────────────────────────────────────────────────

def esc(s):
    """Escape a value for SQL string literal."""
    if s is None:
        return 'null'
    s = str(s).strip()
    if not s or s.lower() in ('none', 'null'):
        return 'null'
    return "'" + s.replace("'", "''") + "'"

def ts(v):
    """Timestamp literal — accepts datetime or string."""
    if v is None:
        return 'null'
    s = str(v).strip()
    if not s:
        return 'null'
    return "'" + s.replace("'", "''") + "'::timestamptz"

def boolish(v):
    """Map various PowerApps representations to SQL boolean."""
    if v is None:
        return 'null'
    s = str(v).strip().lower()
    if s in ('true', 'yes', 'ใช่', '1', 'y'): return 'true'
    if s in ('false', 'no', 'ไม่', '0', 'n'): return 'false'
    return 'null'

def parse_year(reported_at, fallback=2025):
    """Pull YYYY from a timestamp value for JobID prefix."""
    if reported_at is None:
        return fallback
    s = str(reported_at)
    if len(s) >= 4 and s[:4].isdigit():
        return int(s[:4])
    return fallback

# ────────────────────────────────────────────────────────────────────
# Load
# ────────────────────────────────────────────────────────────────────

wb = openpyxl.load_workbook(XLSX, data_only=True)
ws_jobs = wb['ใบแจ้งซ่อม']
ws_insp = wb['ใบตรวจประจำเดือน']

# Header indexes (0-based)
HDR = {
    'reporter_id':    0,   # รหัสพนักงานที่แจ้งเข้ามา
    'reporter_name':  1,   # ชื่อพนักงานที่พบ
    'dept':           2,   # แผนก
    'service_type':   3,   # ประเภทบริการ
    'reported_at':    4,   # เวลาที่แจ้ง
    'repair_type':    5,   # ประเภทการซ่อม
    'item':           6,   # รายการ
    'note':           7,   # หมายเหตุ
    'floor':          8,   # ชั้น
    'zone':           9,   # โซน
    # photos 10..17 — skipped (legacy PowerApps paths)
    'status':         18,  # สถานะ
    'repair_detail':  19,  # วิธีการแก้ไข
    'needs_withdrawal': 20,  # เบิกหรือไม่
    'is_repairable':  21,  # ซ่อมสำเร็จ
    'resolved_at':    22,  # เวลาที่ช่างดำเนินการสำเร็จ
    'data_entry_by':  23,  # ผู้กรอกข้อมูล
    'needs_borrow':   24,  # ยืมหรือไม่
    'needs_return':   25,  # คืนหรือไม่
    'created_at':     26,  # CreatedDate
    'job_id_legacy':  27,  # JobID (may be blank for very old rows)
    'opened_at':      31,  # เวลาที่เปิด
}

# ────────────────────────────────────────────────────────────────────
# Build rpr_jobs INSERTs
# ────────────────────────────────────────────────────────────────────

job_rows = []
year_seq = {}   # year → running sequence for synthetic JobIDs

for r in range(2, ws_jobs.max_row + 1):
    def cell(key):
        idx = HDR.get(key)
        if idx is None: return None
        return ws_jobs.cell(r, idx + 1).value

    reporter_id   = cell('reporter_id')
    reporter_name = cell('reporter_name')
    reported_at   = cell('reported_at')

    # Skip totally blank rows
    if not reporter_id and not reporter_name and not reported_at:
        continue

    # Use legacy JobID if present, else mint one (JOB-YYYY-NNN per year)
    job_id = cell('job_id_legacy')
    if not job_id or not str(job_id).strip():
        yr = parse_year(reported_at)
        year_seq[yr] = year_seq.get(yr, 0) + 1
        job_id = f'JOB-{yr}-{year_seq[yr]:03d}'
    job_id = str(job_id).strip()

    job_rows.append({
        'job_id':           job_id,
        'reporter_id':      str(reporter_id).strip() if reporter_id else None,
        'reporter_name':    str(reporter_name).strip() if reporter_name else None,
        'reporter_dept':    cell('dept'),
        'service_type':     cell('service_type'),
        'repair_type':      cell('repair_type'),
        'item':             cell('item'),
        'note':             cell('note'),
        'floor':            cell('floor'),
        'zone':             cell('zone'),
        'reported_at':      reported_at,
        'opened_at':        cell('opened_at'),
        'status':           cell('status') or 'รอดำเนินการ',
        'repair_detail':    cell('repair_detail'),
        'is_repairable':    cell('is_repairable'),
        'needs_withdrawal': cell('needs_withdrawal'),
        'needs_borrow':     cell('needs_borrow'),
        'needs_return':     cell('needs_return'),
        'data_entry_by':    cell('data_entry_by'),
        'resolved_at':      cell('resolved_at'),
        'created_at':       cell('created_at') or reported_at,
    })

# ────────────────────────────────────────────────────────────────────
# Build rpr_inspections INSERTs
# ────────────────────────────────────────────────────────────────────

INSP = {
    'irid':           0,   # InspectionReportID
    'inspector_id':   1,
    'inspector_name': 2,
    'dept':           3,
    'inspected_at':   4,
    'inspection_type': 5,
    'item':           6,   # รายการที่ตรวจ (free-text often combined w/ detail)
    'floor':          7,
    'zone':           8,
    'detail':         9,
    # photos 10..13 — skipped
    'created_at':     14,
}

insp_rows = []
seq_irid = 0
for r in range(2, ws_insp.max_row + 1):
    def cell(key):
        idx = INSP.get(key)
        if idx is None: return None
        return ws_insp.cell(r, idx + 1).value

    # Skip truly empty rows (xlsx export usually has hundreds of trailing
    # empties). A row is "empty" if it has no inspector + no inspected_at.
    if not cell('inspector_id') and not cell('inspector_name') and not cell('inspected_at'):
        continue

    irid = cell('irid')
    if not irid or not str(irid).strip():
        seq_irid += 1
        irid = f'IRID{seq_irid:06d}'
    irid = str(irid).strip()

    inspector_id = cell('inspector_id')
    if inspector_id is not None:
        # Sheet stores as float "11037.0" — strip trailing .0
        inspector_id = str(inspector_id).split('.')[0]

    insp_rows.append({
        'irid':            irid,
        'inspector_id':    inspector_id,
        'inspector_name':  cell('inspector_name'),
        'dept':            cell('dept'),
        'inspected_at':    cell('inspected_at'),
        'inspection_type': cell('inspection_type'),
        'item':            cell('item'),
        'floor':           cell('floor'),
        'zone':            cell('zone'),
        'detail':          cell('detail'),
        'created_at':      cell('created_at') or cell('inspected_at'),
    })

# ────────────────────────────────────────────────────────────────────
# Write SQL
# ────────────────────────────────────────────────────────────────────

lines = [
    "-- Auto-generated by scripts/repair_import.py",
    f"-- Source: {XLSX.name}",
    f"-- Counts: {len(job_rows)} jobs, {len(insp_rows)} inspections",
    "-- Photos are NOT migrated (legacy PowerApps paths). Re-runnable.",
    "",
    "begin;",
    "",
    "-- ===== rpr_jobs =====",
]

for j in job_rows:
    lines.append(
        f"insert into rpr_jobs (job_id, reporter_id, reporter_name, reporter_dept, "
        f"service_type, repair_type, item, note, floor, zone, "
        f"reported_at, opened_at, status, repair_detail, "
        f"is_repairable, needs_withdrawal, needs_borrow, needs_return, "
        f"data_entry_by, resolved_at, created_at) "
        f"values ({esc(j['job_id'])}, {esc(j['reporter_id'])}, {esc(j['reporter_name'])}, "
        f"{esc(j['reporter_dept'])}, {esc(j['service_type'])}, {esc(j['repair_type'])}, "
        f"{esc(j['item'])}, {esc(j['note'])}, {esc(j['floor'])}, {esc(j['zone'])}, "
        f"{ts(j['reported_at'])}, {ts(j['opened_at'])}, {esc(j['status'])}, "
        f"{esc(j['repair_detail'])}, {boolish(j['is_repairable'])}, "
        f"{boolish(j['needs_withdrawal'])}, {boolish(j['needs_borrow'])}, "
        f"{boolish(j['needs_return'])}, {esc(j['data_entry_by'])}, "
        f"{ts(j['resolved_at'])}, {ts(j['created_at'])}) "
        f"on conflict (job_id) do nothing;"
    )

lines.append("")
lines.append("-- ===== rpr_inspections =====")

def ts_or_default(v):
    """Like ts() but emits DEFAULT for null so NOT NULL columns get now()."""
    if v is None or not str(v).strip():
        return 'default'
    return "'" + str(v).replace("'", "''") + "'::timestamptz"

for i in insp_rows:
    lines.append(
        f"insert into rpr_inspections (irid, inspector_id, inspector_name, dept, "
        f"inspected_at, inspection_type, item, floor, zone, detail, created_at) "
        f"values ({esc(i['irid'])}, {esc(i['inspector_id'])}, {esc(i['inspector_name'])}, "
        f"{esc(i['dept'])}, {ts(i['inspected_at'])}, {esc(i['inspection_type'])}, "
        f"{esc(i['item'])}, {esc(i['floor'])}, {esc(i['zone'])}, {esc(i['detail'])}, "
        f"{ts_or_default(i['created_at'])}) "
        f"on conflict (irid) do nothing;"
    )

lines.append("")
lines.append("commit;")
lines.append("")
lines.append(f"-- Sanity check:")
lines.append(f"-- select count(*) from rpr_jobs;          -- expected ≈ {len(job_rows)}")
lines.append(f"-- select count(*) from rpr_inspections;   -- expected ≈ {len(insp_rows)}")

OUT.write_text('\n'.join(lines), encoding='utf-8')

print(f'Wrote {OUT}')
print(f'  Jobs:        {len(job_rows)}')
print(f'  Inspections: {len(insp_rows)}')
print()
print('Next: run schema_rpr_v1.sql, then this generated SQL in Supabase.')
