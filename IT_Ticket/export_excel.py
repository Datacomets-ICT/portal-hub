"""
Export Supabase tables to Excel (.xlsx)

Usage:
    set SUPABASE_URL=https://rthsmtimvqjnfvgepqpk.supabase.co
    set SUPABASE_KEY=sb_publishable_...   (publishable key OK สำหรับ tickets/worklist)
    python export_excel.py

Output: tickets_export_YYYYMMDD_HHMM.xlsx (3 sheets: tickets, worklist, employees*)
*employees ต้องใช้ secret key ถึงจะ export ได้ (เพราะ RLS)
"""

import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from supabase import create_client

SUPABASE_URL = os.environ.get('SUPABASE_URL')
SUPABASE_KEY = os.environ.get('SUPABASE_KEY')

if not SUPABASE_URL or not SUPABASE_KEY:
    raise SystemExit('ERROR: please set SUPABASE_URL and SUPABASE_KEY environment variables')

sb = create_client(SUPABASE_URL, SUPABASE_KEY)


def fetch_all(table_name, order_col=None, desc=True):
    """Fetch all rows in batches of 1000 (Supabase default limit)."""
    rows = []
    page = 0
    PAGE = 1000
    while True:
        q = sb.table(table_name).select('*').range(page * PAGE, (page + 1) * PAGE - 1)
        if order_col:
            q = q.order(order_col, desc=desc)
        res = q.execute()
        if not res.data:
            break
        rows.extend(res.data)
        if len(res.data) < PAGE:
            break
        page += 1
    return rows


def write_sheet(wb, name, rows, headers=None):
    ws = wb.create_sheet(name)
    if not rows:
        ws.append(['(no data)'])
        return

    cols = headers or list(rows[0].keys())
    ws.append(cols)

    # Header style
    header_fill = PatternFill('solid', fgColor='667EEA')
    header_font = Font(bold=True, color='FFFFFF')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Data rows
    for r in rows:
        ws.append([r.get(c) for c in cols])

    # Auto-width (rough)
    for i, c in enumerate(cols, 1):
        max_len = max(
            [len(str(c))] + [len(str(r.get(c) or '')) for r in rows[:200]]
        )
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 40)

    # Freeze header
    ws.freeze_panes = 'A2'


print('Fetching tickets...')
tickets = fetch_all('tickets', order_col='created_at', desc=True)
print(f'  {len(tickets)} rows')

print('Fetching worklist...')
worklist = fetch_all('worklist', order_col='id', desc=False)
print(f'  {len(worklist)} rows')

print('Fetching employees...')
try:
    employees = fetch_all('employees', order_col='employee_id', desc=False)
    print(f'  {len(employees)} rows')
except Exception as e:
    print(f'  skipped (need secret key): {e}')
    employees = []

# Build workbook
wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

write_sheet(wb, 'tickets', tickets)
write_sheet(wb, 'worklist', worklist)
if employees:
    write_sheet(wb, 'employees', employees)

stamp = datetime.now().strftime('%Y%m%d_%H%M')
filename = f'tickets_export_{stamp}.xlsx'
wb.save(filename)
print(f'\nDone -> {filename}')
