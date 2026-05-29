"""
Migration v2: request (7).xlsx -> Supabase (overwrite existing data)

Usage:
    set SUPABASE_URL=https://rthsmtimvqjnfvgepqpk.supabase.co
    set SUPABASE_SERVICE_KEY=sb_secret_...
    python migrate_v2.py
"""

import os
from datetime import datetime

import openpyxl
from supabase import create_client

EXCEL_FILE = os.path.join(os.path.dirname(__file__), 'requests (7).xlsx')

SUPABASE_URL = os.environ.get('SUPABASE_URL')
SUPABASE_KEY = os.environ.get('SUPABASE_SERVICE_KEY')

if not SUPABASE_URL or not SUPABASE_KEY:
    raise SystemExit('ERROR: set SUPABASE_URL and SUPABASE_SERVICE_KEY')

sb = create_client(SUPABASE_URL, SUPABASE_KEY)
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)


def cell_str(v):
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip() or None


def clean(s):
    """Remove null bytes that Postgres rejects."""
    return s.replace('\x00', '') if isinstance(s, str) else s


# ---------- Admin list ----------
print('Reading Admin sheet...')
admin_ws = wb['Admin']
active_admins = set()
all_admins = {}
for row in admin_ws.iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    name = str(row[0]).strip()
    active = str(row[1] or '').strip().lower() == 'true'
    # Extract ID part (e.g. "IT03_วี" -> "IT03")
    admin_id = name.split('_')[0] if '_' in name else name
    nickname = name.split('_')[1] if '_' in name else name
    all_admins[admin_id] = {'nickname': nickname, 'active': active}
    if active:
        active_admins.add(admin_id)

print(f'  Total admins: {len(all_admins)}, Active: {len(active_admins)}')
print(f'  Active: {active_admins}')


# ---------- Employees ----------
print('Reading Employee sheet...')
emp_ws = wb['Employee']
employees = []
for row in emp_ws.iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    emp_id = cell_str(row[0])
    resigned = None
    if row[10]:
        if isinstance(row[10], datetime):
            resigned = row[10].strftime('%Y-%m-%d')
        else:
            resigned = str(row[10])[:10] if row[10] else None

    is_admin = emp_id in active_admins or emp_id == 'Admin'

    employees.append({
        'employee_id': emp_id,
        'password':    cell_str(row[1]) or emp_id,  # default password = emp_id
        'company':     cell_str(row[2]),
        'first_name':  cell_str(row[3]),
        'last_name':   cell_str(row[4]),
        'department':  cell_str(row[5]),
        'section':     cell_str(row[6]),
        'position':    cell_str(row[7]),
        'nickname':    cell_str(row[8]),
        'email':       cell_str(row[11]) if len(row) > 11 else None,
        'phone':       cell_str(row[12]) if len(row) > 12 else None,
        'is_admin':    is_admin,
        'is_approved':  True,
        'resigned_date': resigned,
    })

# Add IT admin accounts that might not be in Employee sheet
for admin_id, info in all_admins.items():
    if not any(e['employee_id'] == admin_id for e in employees):
        employees.append({
            'employee_id': admin_id,
            'password': '1234',
            'company': '',
            'first_name': admin_id,
            'last_name': '',
            'department': 'IT',
            'section': 'IT',
            'position': 'IT Support',
            'nickname': info['nickname'],
            'is_admin': info['active'],
            'is_approved': True,
            'resigned_date': None,
        })

print(f'  Total employees: {len(employees)}')
print(f'  Upserting...')

# Clear and re-insert employees
sb.table('employees').delete().gte('employee_id', '').execute()
BATCH = 200
for i in range(0, len(employees), BATCH):
    batch = [{k: clean(v) if isinstance(v, str) else v for k, v in e.items()} for e in employees[i:i+BATCH]]
    sb.table('employees').upsert(batch).execute()
    print(f'    {i + len(batch)}/{len(employees)}')


# ---------- Worklist ----------
print('Reading Worklist sheet...')
wl_ws = wb['Worklist']
worklist = []
for row in wl_ws.iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    worklist.append({
        'job_type':   cell_str(row[0]),
        'issue_type': cell_str(row[1]),
        'symptom':    cell_str(row[2]),
    })

print(f'  {len(worklist)} rows, replacing...')
sb.table('worklist').delete().gte('id', 0).execute()
if worklist:
    sb.table('worklist').insert(worklist).execute()


# ---------- Tickets ----------
print('Reading requests sheet...')
req_ws = wb['requests']
tickets = []
for row in req_ws.iter_rows(min_row=2, values_only=True):
    if not row[1]:
        continue
    created = row[2] if isinstance(row[2], datetime) else None
    tickets.append({
        'ticket_no':     cell_str(row[1]),
        'created_at':    created.isoformat() if created else None,
        'employee_id':   cell_str(row[3]),
        'employee_name': cell_str(row[4]),
        'plant':         cell_str(row[5]),
        'job_type':      cell_str(row[7]),
        'issue_type':    cell_str(row[8]),
        'symptom':       cell_str(row[9]),
        'detail':        cell_str(row[10]),
        'request':       cell_str(row[13]),
        'vnc_number':    cell_str(row[14]),
        'phone':         cell_str(row[15]),
        'status':        cell_str(row[17]) or 'เปิด Ticket',
        'reply':         cell_str(row[18]),
        'admin':         cell_str(row[20]),
        'location':      cell_str(row[21]) if len(row) > 21 else None,
    })

# Dedupe by ticket_no
seen = {}
for t in tickets:
    seen[t['ticket_no']] = t
tickets = list(seen.values())

print(f'  {len(tickets)} tickets (deduped), upserting...')

# Clear and re-insert
sb.table('tickets').delete().gte('key', 0).execute()
for i in range(0, len(tickets), BATCH):
    batch = [{k: clean(v) if isinstance(v, str) else v for k, v in t.items()} for t in tickets[i:i+BATCH]]
    sb.table('tickets').insert(batch).execute()
    print(f'    {i + len(batch)}/{len(tickets)}')

print('\nDone! Migration complete.')
print(f'  Employees: {len(employees)}')
print(f'  Worklist:  {len(worklist)}')
print(f'  Tickets:   {len(tickets)}')
print(f'  Admins:    {len(active_admins)} active')
