"""
One-time migration: requests (6).xlsx -> Supabase

Usage:
    pip install openpyxl supabase
    set SUPABASE_URL=https://YOUR_PROJECT.supabase.co
    set SUPABASE_SERVICE_KEY=eyJ...    (Settings -> API -> service_role key)
    python migrate.py

NOTE: Use the SERVICE ROLE key (not anon) so we can bypass RLS to insert into employees.
Run schema.sql in Supabase SQL Editor BEFORE running this script.
"""

import os
from datetime import datetime

import openpyxl
from supabase import create_client

EXCEL_FILE = os.path.join(os.path.dirname(__file__), 'requests (6).xlsx')

SUPABASE_URL = os.environ.get('SUPABASE_URL')
SUPABASE_SERVICE_KEY = os.environ.get('SUPABASE_SERVICE_KEY')

if not SUPABASE_URL or not SUPABASE_SERVICE_KEY:
    raise SystemExit('ERROR: please set SUPABASE_URL and SUPABASE_SERVICE_KEY environment variables')

sb = create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)


def cell_str(v):
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip() or None


# ---------- Employees ----------
print('Reading Employee sheet...')
emp_ws = wb['Employee']
employees = []
for row in emp_ws.iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    employees.append({
        'employee_id': cell_str(row[0]),
        'password':    cell_str(row[1]) or '',
        'company':     cell_str(row[2]),
        'first_name':  cell_str(row[3]),
        'last_name':   cell_str(row[4]),
        'department':  cell_str(row[5]),
        'section':     cell_str(row[6]),
        'position':    cell_str(row[7]),
        'nickname':    cell_str(row[8]),
        'email':       cell_str(row[11]) if len(row) > 11 else None,
        'phone':       cell_str(row[12]) if len(row) > 12 else None,
    })

print(f'  Upserting {len(employees)} employees...')
if employees:
    sb.table('employees').upsert(employees).execute()


# ---------- Worklist ----------
print('Reading Worklist sheet...')
wl_ws = wb['Worklist']
worklist = []
for row in wl_ws.iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    worklist.append({
        'job_type':   cell_str(row[0]),
        'issue_type': cell_str(row[1]),
        'symptom':    cell_str(row[2]),
    })

print(f'  Replacing {len(worklist)} worklist rows...')
sb.table('worklist').delete().gte('id', 0).execute()
if worklist:
    sb.table('worklist').insert(worklist).execute()


# ---------- Tickets ----------
print('Reading requests sheet...')
req_ws = wb['requests']
tickets = []
for row in req_ws.iter_rows(min_row=2, values_only=True):
    if not row[1]:
        continue
    created = row[2] if isinstance(row[2], datetime) else None
    tickets.append({
        'ticket_no':     cell_str(row[1]),
        'created_at':    created.isoformat() if created else None,
        'employee_id':   cell_str(row[3]),
        'employee_name': cell_str(row[4]),
        'plant':         cell_str(row[5]),
        'job_type':      cell_str(row[7]),
        'issue_type':    cell_str(row[8]),
        'symptom':       cell_str(row[9]),
        'detail':        cell_str(row[10]),
        'request':       cell_str(row[13]),
        'vnc_number':    cell_str(row[14]),
        'phone':         cell_str(row[15]),
        'status':        cell_str(row[17]) or 'เปิด Ticket',
        'reply':         cell_str(row[18]),
        'admin':         cell_str(row[20]),
        'location':      cell_str(row[21]) if len(row) > 21 else None,
    })

# Dedupe by ticket_no (keep last occurrence) — Excel may have duplicates
seen = {}
for t in tickets:
    seen[t['ticket_no']] = t
tickets = list(seen.values())

print(f'  Upserting {len(tickets)} tickets (deduped)...')
# Insert in batches of 500 to avoid payload limits
BATCH = 500
for i in range(0, len(tickets), BATCH):
    sb.table('tickets').upsert(tickets[i:i + BATCH], on_conflict='ticket_no').execute()

print('Done.')
